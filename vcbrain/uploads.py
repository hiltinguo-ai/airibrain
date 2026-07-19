"""Parse founder file uploads into pipeline inputs.

Accepted formats:
  .pdf                 → pitch deck text (pypdf extraction)
  .md / .txt           → deck text as-is
  .csv                 → sniffed: month/revenue table → revenue series;
                         key,value rows → headline metrics
  .xlsx / .xlsm        → every sheet sniffed the same way (openpyxl)
  .json                → numeric-valued object → metrics; string-valued → founder Q&A

Returns the same flat payload the webapp's paste path uses:
  {"deck": str, "metrics": json-str, "revenue": csv-str, "qa": json-str}
"""

from __future__ import annotations

import csv
import io
import json
import logging

logger = logging.getLogger("vcbrain.uploads")


def _pdf_text(data: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    return "\n".join((page.extract_text() or "") for page in reader.pages)


def _num(v) -> float | None:
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").replace("$", "").replace("%", "").strip())
    except (ValueError, AttributeError):
        return None


def _ingest_table(rows: list[list], metrics: dict, revenue: list[dict]) -> None:
    """Classify a rectangular table as a revenue series or a metrics sheet."""
    rows = [r for r in rows if any(c not in (None, "") for c in r)]
    if not rows:
        return
    header = [str(c or "").strip().lower() for c in rows[0]]
    if "month" in header and "revenue" in header:
        mi, ri = header.index("month"), header.index("revenue")
        for r in rows[1:]:
            val = _num(r[ri]) if len(r) > ri else None
            if val is not None and len(r) > mi and r[mi] not in (None, ""):
                revenue.append({"month": str(r[mi]).strip(), "revenue": val})
        return
    # key/value metrics: two usable columns, numeric second column
    for r in rows:
        if len(r) >= 2 and r[0] not in (None, ""):
            val = _num(r[1])
            if val is not None:
                key = str(r[0]).strip().lower().replace(" ", "_")
                metrics[key] = val


def _rows_from_csv(data: bytes) -> list[list]:
    text = data.decode("utf-8-sig", errors="replace")
    return [row for row in csv.reader(io.StringIO(text))]


def _rows_from_xlsx(data: bytes) -> list[list[list]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        sheets.append([list(row) for row in ws.iter_rows(values_only=True)])
    wb.close()
    return sheets


def parse_files(files) -> dict:
    """`files` is an iterable of Werkzeug FileStorage (or anything with
    .filename and .read()). Unknown extensions are skipped with a log line."""
    deck_parts: list[str] = []
    metrics: dict = {}
    revenue: list[dict] = []
    qa: dict = {}

    for f in files:
        name = (getattr(f, "filename", "") or "").lower()
        if not name:
            continue
        data = f.read()
        try:
            if name.endswith(".pdf"):
                deck_parts.append(_pdf_text(data))
            elif name.endswith((".md", ".txt")):
                deck_parts.append(data.decode("utf-8", errors="replace"))
            elif name.endswith(".csv"):
                _ingest_table(_rows_from_csv(data), metrics, revenue)
            elif name.endswith((".xlsx", ".xlsm", ".xltx")):
                for sheet in _rows_from_xlsx(data):
                    _ingest_table(sheet, metrics, revenue)
            elif name.endswith(".json"):
                obj = json.loads(data.decode("utf-8", errors="replace"))
                if isinstance(obj, dict):
                    nums = {k: v for k, v in obj.items() if _num(v) is not None
                            and not isinstance(v, str)}
                    if nums and len(nums) >= len(obj) / 2:
                        metrics.update({k: float(v) for k, v in nums.items()})
                    else:
                        qa.update({str(k): str(v) for k, v in obj.items()})
            else:
                logger.warning("skipping unsupported upload: %s", name)
        except Exception:
            logger.exception("failed to parse upload %s — skipping", name)

    revenue_csv = ""
    if revenue:
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(["month", "revenue"])
        for p in revenue:
            w.writerow([p["month"], p["revenue"]])
        revenue_csv = out.getvalue()

    return {
        "deck": "\n\n".join(x for x in deck_parts if x.strip()),
        "metrics": json.dumps(metrics) if metrics else "",
        "revenue": revenue_csv,
        "qa": json.dumps(qa) if qa else "",
    }
